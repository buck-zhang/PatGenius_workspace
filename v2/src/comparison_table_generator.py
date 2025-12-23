#!/usr/bin/env python3
"""
構成対比表生成モジュール - v2.1.2

進歩性判断エンジンの結果から、弁理士向けの構成対比表を自動生成する。

機能:
1. X文献との構成対比表
2. Y文献（主引例・副引例）との構成対比表
3. Excel形式での出力（色分け、フィルター付き）
4. Markdown形式での出力（オプション）
"""

import json
import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ComparisonTableGenerator:
    """構成対比表生成クラス"""

    def __init__(self):
        """初期化"""
        # セルのスタイル定義
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)

        self.disclosed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 薄い緑
        self.not_disclosed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 薄い赤

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')

    def generate_comparison_table(
        self,
        assessment_summary_path: str,
        comparison_results_dir: str,
        base_structure_path: str,
        output_path: str,
        format: str = "excel"
    ) -> str:
        """
        構成対比表を生成

        Args:
            assessment_summary_path: 進歩性判断サマリーJSONのパス
            comparison_results_dir: 個別構成対比結果のディレクトリ
            base_structure_path: 本願特許の構成要素JSONパス
            output_path: 出力ファイルパス
            format: 出力形式（"excel" または "markdown"）

        Returns:
            生成されたファイルのパス
        """
        print(f"\n{'='*80}")
        print(f"構成対比表の生成を開始")
        print(f"{'='*80}\n")

        # 1. データ読み込み
        print("Step 1: データ読み込み中...")
        summary = self._load_assessment_results(assessment_summary_path)
        comparisons = self._load_comparison_results(comparison_results_dir)
        base_structure = self._load_base_structure(base_structure_path)

        print(f"  ✓ 本願特許の構成要素: {len(base_structure['構成要件'])}個")
        print(f"  ✓ 構成対比結果: {len(comparisons)}件")

        # 2. X文献の選択
        print("\nStep 2: X文献の選択...")
        x_reference = summary['x_references']['patents'][0] if summary['x_references']['patents'] else None
        if x_reference:
            print(f"  ✓ X文献: {x_reference}")
        else:
            print(f"  ⚠ X文献なし")

        # 3. 代表的なY文献の選択
        print("\nStep 3: 代表的なY文献の選択...")
        y_reference = self._select_representative_y_reference(
            summary['y_references']['combinations']
        )
        if y_reference:
            print(f"  ✓ Y文献（主引例）: {y_reference['primary_reference']}")
            print(f"  ✓ Y文献（副引例）: {', '.join(y_reference['secondary_references'])}")
        else:
            print(f"  ⚠ Y文献なし")

        # 4. 表データの構築
        print("\nStep 4: 表データの構築中...")
        table_data = self._build_table_data(
            base_structure['構成要件'],
            x_reference,
            y_reference,
            comparisons
        )
        print(f"  ✓ {len(table_data)}行のデータを構築")

        # 5. 出力
        print(f"\nStep 5: {format}形式で出力中...")
        if format == "excel":
            output_file = self._export_to_excel(table_data, output_path, summary)
        else:
            output_file = self._export_to_markdown(table_data, output_path)

        print(f"\n{'='*80}")
        print(f"✅ 構成対比表の生成完了")
        print(f"{'='*80}")
        print(f"出力ファイル: {output_file}\n")

        return output_file

    def _load_assessment_results(self, summary_path: str) -> Dict:
        """進歩性判断サマリーの読み込み"""
        with open(summary_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _load_comparison_results(self, results_dir: str) -> Dict[str, Dict]:
        """個別構成対比結果の読み込み"""
        results_dir = Path(results_dir)
        comparisons = {}

        for file in results_dir.glob("comparison_*.json"):
            with open(file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                patent_id = data['target_patent_id']
                comparisons[patent_id] = data

        return comparisons

    def _load_base_structure(self, structure_path: str) -> Dict:
        """本願特許の構成要素の読み込み"""
        with open(structure_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _select_representative_y_reference(
        self,
        y_combinations: List[Dict]
    ) -> Optional[Dict]:
        """
        代表的なY文献組み合わせを選択

        選択基準:
        1. 主引例の開示要素数が最多
        2. 同数の場合、副引例の開示要素数が最少
        3. それでも同じ場合、主引例の特許番号が最も古い

        Args:
            y_combinations: Y文献の組み合わせリスト

        Returns:
            代表的なY文献組み合わせ（なければNone）
        """
        if not y_combinations:
            return None

        # 各組み合わせをスコアリング
        scored_combinations = []

        for combo in y_combinations:
            primary_id = combo['primary_reference']
            primary_coverage = len(combo['coverage'][primary_id])

            # 副引例の合計カバレッジ
            secondary_coverage = sum(
                len(combo['coverage'][sec_id])
                for sec_id in combo['secondary_references']
            )

            # スコア: (主引例カバレッジ, -副引例カバレッジ, -主引例ID)
            # 主引例が多く、副引例が少なく、IDが古いほど優先
            score = (
                primary_coverage,
                -secondary_coverage,
                -int(''.join(filter(str.isdigit, primary_id)))  # 特許番号の数値部分
            )

            scored_combinations.append((score, combo))

        # スコアでソートして最良のものを選択
        scored_combinations.sort(key=lambda x: x[0], reverse=True)

        return scored_combinations[0][1]

    def _build_table_data(
        self,
        base_elements: List[Dict],
        x_reference: Optional[str],
        y_reference: Optional[Dict],
        comparisons: Dict[str, Dict]
    ) -> List[Dict]:
        """
        表データの構築

        Args:
            base_elements: 本願特許の構成要素リスト
            x_reference: X文献の特許ID
            y_reference: Y文献の組み合わせ情報
            comparisons: 個別構成対比結果

        Returns:
            表データのリスト（各行が辞書）
        """
        table_rows = []

        for element in base_elements:
            element_id = element['構成要素番号']
            element_desc = element.get('構成要素の簡単説明', element['構成要素'])

            row = {
                'element_id': element_id,
                'element_description': element_desc,
                'is_independent': element.get('is_independent', False),
            }

            # X文献の情報を取得
            if x_reference and x_reference in comparisons:
                x_data = self._get_element_evidence(
                    comparisons[x_reference],
                    element_id
                )
                row['x_reference_id'] = x_reference
                row['x_disclosed'] = x_data['disclosed']
                row['x_content'] = x_data['content']
                row['x_location'] = x_data['location']
            else:
                row['x_reference_id'] = '-'
                row['x_disclosed'] = False
                row['x_content'] = '-'
                row['x_location'] = '-'

            # Y文献（主引例）の情報を取得
            if y_reference:
                primary_id = y_reference['primary_reference']
                if primary_id in comparisons:
                    y_primary_data = self._get_element_evidence(
                        comparisons[primary_id],
                        element_id
                    )
                    row['y_primary_id'] = primary_id
                    row['y_primary_disclosed'] = y_primary_data['disclosed']
                    row['y_primary_content'] = y_primary_data['content']
                    row['y_primary_location'] = y_primary_data['location']
                else:
                    row['y_primary_id'] = primary_id
                    row['y_primary_disclosed'] = False
                    row['y_primary_content'] = '-'
                    row['y_primary_location'] = '-'

                # Y文献（副引例）の情報を取得
                secondary_ids = y_reference['secondary_references']
                if secondary_ids and len(secondary_ids) > 0:
                    # 最初の副引例
                    sec_id = secondary_ids[0]
                    if sec_id in comparisons:
                        y_sec_data = self._get_element_evidence(
                            comparisons[sec_id],
                            element_id
                        )
                        row['y_secondary_id'] = sec_id
                        row['y_secondary_disclosed'] = y_sec_data['disclosed']
                        row['y_secondary_content'] = y_sec_data['content']
                        row['y_secondary_location'] = y_sec_data['location']
                    else:
                        row['y_secondary_id'] = sec_id
                        row['y_secondary_disclosed'] = False
                        row['y_secondary_content'] = '-'
                        row['y_secondary_location'] = '-'
                else:
                    row['y_secondary_id'] = '-'
                    row['y_secondary_disclosed'] = False
                    row['y_secondary_content'] = '-'
                    row['y_secondary_location'] = '-'
            else:
                row['y_primary_id'] = '-'
                row['y_primary_disclosed'] = False
                row['y_primary_content'] = '-'
                row['y_primary_location'] = '-'
                row['y_secondary_id'] = '-'
                row['y_secondary_disclosed'] = False
                row['y_secondary_content'] = '-'
                row['y_secondary_location'] = '-'

            table_rows.append(row)

        return table_rows

    def _get_element_evidence(
        self,
        comparison_result: Dict,
        element_id: str
    ) -> Dict:
        """
        特定要素の開示証拠を取得

        Args:
            comparison_result: 構成対比結果
            element_id: 要素ID

        Returns:
            開示情報の辞書
        """
        for elem in comparison_result['element_comparisons']:
            if elem['element_id'] == element_id:
                evidence = elem.get('evidence', {})
                return {
                    'disclosed': elem.get('is_disclosed', False),
                    'content': evidence.get('quoted_text', '-') if elem.get('is_disclosed') else '-',
                    'location': ', '.join(evidence.get('locations', [])) if elem.get('is_disclosed') else '-'
                }

        # 要素が見つからない場合
        return {
            'disclosed': False,
            'content': '-',
            'location': '-'
        }

    def _export_to_excel(
        self,
        table_data: List[Dict],
        output_path: str,
        summary: Dict
    ) -> str:
        """
        Excel形式で出力

        Args:
            table_data: 表データ
            output_path: 出力ファイルパス
            summary: 進歩性判断サマリー

        Returns:
            出力ファイルパス
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "構成対比表"

        # ヘッダー行を追加
        headers = [
            "構成要素\n番号",
            "構成要素概要",
            f"X文献\n{table_data[0]['x_reference_id']}" if table_data[0]['x_reference_id'] != '-' else "X文献",
            "X文献の\n記載内容",
            "X文献の\n記載箇所",
            f"Y文献（主引例）\n{table_data[0]['y_primary_id']}" if table_data[0]['y_primary_id'] != '-' else "Y文献（主引例）",
            "Y文献（主引例）の\n記載内容",
            "Y文献（主引例）の\n記載箇所",
            f"Y文献（副引例）\n{table_data[0]['y_secondary_id']}" if table_data[0]['y_secondary_id'] != '-' else "Y文献（副引例）",
            "Y文献（副引例）の\n記載内容",
            "Y文献（副引例）の\n記載箇所"
        ]

        # ヘッダー行を書き込み
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border

        # データ行を書き込み
        for row_num, row_data in enumerate(table_data, 2):
            # 構成要素番号
            cell = ws.cell(row=row_num, column=1)
            cell.value = row_data['element_id']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            if row_data['is_independent']:
                cell.font = Font(bold=True)

            # 構成要素概要
            cell = ws.cell(row=row_num, column=2)
            cell.value = row_data['element_description']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # X文献の開示状況
            cell = ws.cell(row=row_num, column=3)
            cell.value = "✓" if row_data['x_disclosed'] else "－"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            if row_data['x_disclosed']:
                cell.fill = self.disclosed_fill
            else:
                cell.fill = self.not_disclosed_fill

            # X文献の記載内容
            cell = ws.cell(row=row_num, column=4)
            cell.value = row_data['x_content']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # X文献の記載箇所
            cell = ws.cell(row=row_num, column=5)
            cell.value = row_data['x_location']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # Y文献（主引例）の開示状況
            cell = ws.cell(row=row_num, column=6)
            cell.value = "✓" if row_data['y_primary_disclosed'] else "－"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            if row_data['y_primary_disclosed']:
                cell.fill = self.disclosed_fill
            else:
                cell.fill = self.not_disclosed_fill

            # Y文献（主引例）の記載内容
            cell = ws.cell(row=row_num, column=7)
            cell.value = row_data['y_primary_content']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # Y文献（主引例）の記載箇所
            cell = ws.cell(row=row_num, column=8)
            cell.value = row_data['y_primary_location']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # Y文献（副引例）の開示状況
            cell = ws.cell(row=row_num, column=9)
            cell.value = "✓" if row_data['y_secondary_disclosed'] else "－"
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
            if row_data['y_secondary_disclosed']:
                cell.fill = self.disclosed_fill
            else:
                cell.fill = self.not_disclosed_fill

            # Y文献（副引例）の記載内容
            cell = ws.cell(row=row_num, column=10)
            cell.value = row_data['y_secondary_content']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

            # Y文献（副引例）の記載箇所
            cell = ws.cell(row=row_num, column=11)
            cell.value = row_data['y_secondary_location']
            cell.alignment = self.wrap_alignment
            cell.border = self.border

        # 列幅の調整
        ws.column_dimensions['A'].width = 12  # 構成要素番号
        ws.column_dimensions['B'].width = 40  # 構成要素概要
        ws.column_dimensions['C'].width = 15  # X文献開示状況
        ws.column_dimensions['D'].width = 50  # X文献記載内容
        ws.column_dimensions['E'].width = 20  # X文献記載箇所
        ws.column_dimensions['F'].width = 15  # Y主引例開示状況
        ws.column_dimensions['G'].width = 50  # Y主引例記載内容
        ws.column_dimensions['H'].width = 20  # Y主引例記載箇所
        ws.column_dimensions['I'].width = 15  # Y副引例開示状況
        ws.column_dimensions['J'].width = 50  # Y副引例記載内容
        ws.column_dimensions['K'].width = 20  # Y副引例記載箇所

        # 行の高さ調整
        ws.row_dimensions[1].height = 40  # ヘッダー行

        # フィルターを追加
        ws.auto_filter.ref = ws.dimensions

        # ウィンドウ枠の固定（ヘッダー行）
        ws.freeze_panes = "A2"

        # ファイル保存
        output_path = Path(output_path)
        if output_path.suffix != '.xlsx':
            output_path = output_path.with_suffix('.xlsx')

        wb.save(output_path)

        return str(output_path)

    def _export_to_markdown(
        self,
        table_data: List[Dict],
        output_path: str
    ) -> str:
        """
        Markdown形式で出力

        Args:
            table_data: 表データ
            output_path: 出力ファイルパス

        Returns:
            出力ファイルパス
        """
        output_path = Path(output_path)
        if output_path.suffix != '.md':
            output_path = output_path.with_suffix('.md')

        with open(output_path, 'w', encoding='utf-8') as f:
            # タイトル
            f.write("# 構成対比表\n\n")
            f.write(f"**生成日時**: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}\n\n")

            # 表のヘッダー
            x_ref = table_data[0]['x_reference_id']
            y_primary = table_data[0]['y_primary_id']
            y_secondary = table_data[0]['y_secondary_id']

            f.write("| 要素番号 | 要素概要 | X文献<br>")
            f.write(f"{x_ref} | 記載内容 | 記載箇所 | Y文献（主引例）<br>")
            f.write(f"{y_primary} | 記載内容 | 記載箇所 | Y文献（副引例）<br>")
            f.write(f"{y_secondary} | 記載内容 | 記載箇所 |\n")

            f.write("|" + "---|" * 11 + "\n")

            # データ行
            for row in table_data:
                elem_id = f"**{row['element_id']}**" if row['is_independent'] else row['element_id']
                x_status = "✅" if row['x_disclosed'] else "❌"
                y_primary_status = "✅" if row['y_primary_disclosed'] else "❌"
                y_secondary_status = "✅" if row['y_secondary_disclosed'] else "❌"

                # 内容を短縮（最初の100文字）
                x_content = row['x_content'][:100] + "..." if len(row['x_content']) > 100 else row['x_content']
                y_primary_content = row['y_primary_content'][:100] + "..." if len(row['y_primary_content']) > 100 else row['y_primary_content']
                y_secondary_content = row['y_secondary_content'][:100] + "..." if len(row['y_secondary_content']) > 100 else row['y_secondary_content']

                f.write(f"| {elem_id} | {row['element_description'][:50]} | {x_status} | ")
                f.write(f"{x_content} | {row['x_location']} | {y_primary_status} | ")
                f.write(f"{y_primary_content} | {row['y_primary_location']} | {y_secondary_status} | ")
                f.write(f"{y_secondary_content} | {row['y_secondary_location']} |\n")

            f.write("\n")
            f.write("凡例: ✅ = 開示あり, ❌ = 開示なし\n")

        return str(output_path)


if __name__ == "__main__":
    # テスト実行
    import sys

    if len(sys.argv) < 5:
        print("使用方法: python comparison_table_generator.py <summary_json> <comparisons_dir> <structure_json> <output_path> [format]")
        sys.exit(1)

    summary_path = sys.argv[1]
    comparisons_dir = sys.argv[2]
    structure_path = sys.argv[3]
    output_path = sys.argv[4]
    format_type = sys.argv[5] if len(sys.argv) > 5 else "excel"

    generator = ComparisonTableGenerator()
    result = generator.generate_comparison_table(
        assessment_summary_path=summary_path,
        comparison_results_dir=comparisons_dir,
        base_structure_path=structure_path,
        output_path=output_path,
        format=format_type
    )

    print(f"✅ 構成対比表を生成しました: {result}")
